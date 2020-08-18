import openpyxl
import random
from tqdm import tqdm

from card import Card

def loadCardList(filename):
    """
    カードリストが記述されたxlsxファイルを読み込む。
    基本土地枠に基本土地でないカードが入る場合、その内容が記述されたファイルを読み込む。
    レアリティごとのリスト、全カードリストを作成する。

    Parameters
    ----------
    filename : str
        読み込む対象のxlsxファイル名

    Returns
    -------
    all : list(Card)
        全てのカードリスト
    cList : list(Card)
        コモンのカードリスト
    uList : list(Card)
        アンコモンのカードリスト
    rList : list(Card)
        レアのカードリスト
    mList : list(Card)
        神話のカードリスト
    lList : list(Card)
        土地のカードリスト
    """
    all, cList, uList, rList, mList, lList = [], [], [], [], [], []
    print("\n----- Start Loading CardList -----")
    # 基本でない土地枠の読み込み
    try :
        with open("notBasicLands.txt") as f:
            print("loading notBasicLands.")
            l = [s.strip() for s in f.readlines()]
    except FileNotFoundError as e:
        print("only basic land.")
        l = []

    # カードリスト読み込み
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError as e:
        print("xlsx file does not found.")
        return all, cList, uList, rList, mList, lList
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]
    for i in tqdm(range(2, sheet.max_row+1)):
        name_jp = sheet.cell(row=i,column=1).value
        name_en = sheet.cell(row=i,column=2).value
        rarity = sheet.cell(row=i,column=3).value
        card = Card(name_jp, name_en, rarity)
        all.append(card)
        if card.rarity == "C":
            if card.name_en in l:
                lList.append(card)
            else:
                cList.append(card)
        elif card.rarity == "U":
            uList.append(card)
        elif card.rarity == "R":
            rList.append(card)
        elif card.rarity == "M":
            mList.append(card)
        elif card.rarity == "L":
            lList.append(card)
    print("----- Finish Loading CardList -----")

    return all, cList, uList, rList, mList, lList

def pickCard(cardList, rand):
    """
    カードリストからランダムなカードを取得する。

    Parameters
    ----------
    cardList : list(Card)
        対象のカードリスト。
    rand : int
        乱数値

    Returns
    -------
    card : Card
        ランダムに選ばれたカード。
    """
    try:
        card = cardList[rand]
    except IndexError as e:
        print(e)
        card = None
    return card

# pack開封モード
def openPack(all, cList, uList, rList, mList, lList, rand_foil, rand_mythic):
    """
    1パック開封するゲームモードを開始する

    Parameters
    ----------
    all : list(Card)
        全てのカードリスト
    cList : list(Card)
        コモンのカードリスト
    uList : list(Card)
        アンコモンのカードリスト
    rList : list(Card)
        レアのカードリスト
    mList : list(Card)
        神話のカードリスト
    lList : list(Card)
        土地のカードリスト
    rand_foil : int
        フォイルの抽選乱数
    rand_mythic : int
        神話レアの抽選乱数

    Returns
    -------
    pack : list(Card)
        作成されたパックの内容
    """

    pack = []

    # Foilの抽選
    # 1/3の確率でFoilが封入
    if rand_foil == 0:
        foilCard = pickCard(all, random.randrange(len(all)))
        if foilCard != None:
            foilCard.setFoiled(True)
            pack.append(foilCard)

    # Mythicの抽選
    # 1/8の確率でMythic, そうでなければRare
    if rand_mythic == 0:
        rareCard = pickCard(mList, random.randrange(len(mList)))
    else:
        rareCard = pickCard(rList, random.randrange(len(rList)))
    if rareCard != None:
        pack.append(rareCard)

    # Uncommonの抽選
    # 1パックに3枚
    uc_i = 0
    while uc_i != 3:
        ucCard = pickCard(uList, random.randrange(len(uList)))
        # 同じカードが抽選された場合はやり直し
        if ucCard in pack:
            continue
        if ucCard != None:
            pack.append(ucCard)
        uc_i+=1

    # Commonの抽選
    # レア枠、UC枠、Foil枠、土地枠以外全て
    while len(pack) != 14:
        cCard = pickCard(cList, random.randrange(len(cList)))
        # 同じカードが抽選された場合はやり直し
        if cCard in pack:
            continue
        if cCard != None:
            pack.append(cCard)

    # Landの抽選
    landCard = pickCard(lList, random.randrange(len(lList)))
    if landCard != None:
        pack.append(landCard)

    if len(pack) != 15:
        print("Failed to create pack!")
        return []
    else:
        return pack

def printCards(cardList, title):
    """
    対象のリストに含まれるカード情報、その題名を出力する
    Parameters
    ----------
    cardList : list(Card)
        対象のカードリスト
    title : str
        出力する題名
    """
    print("\n" + title)
    print("--------------------------------------------------------------------------------------")
    for card in cardList:
        card.print()
    print("--------------------------------------------------------------------------------------")

def main():
    allCardList, commonCardList, uncommonCardList, rareCardList, mythicCardList, landCardList = loadCardList("M21.xlsx")

    print("\nWelcome to Open Pack Simulator!")
    while True:
        print("\nMenu List-----------------------------------------------------------------------------")
        print("1:open a pack")
        print("2:show all common cards")
        print("3:show all uncommon cards")
        print("4:show all rare cards")
        print("5:show all mythic cards")
        print("6:show all land cards")
        print("7:show all cards")
        print("0:exit this program")
        print("--------------------------------------------------------------------------------------")
        print("input number>", end="")
        try:
            check = int(input())
        except ValueError as e:
            print("\nError: ", end="")
            print(e)
            print("Retry input number.")
            continue
        if check == 1:
            pack = openPack(allCardList, commonCardList, uncommonCardList, rareCardList, mythicCardList, landCardList, random.randrange(3), random.randrange(8))
            printCards(pack, "Open a Pack!")
        elif check == 2:
            printCards(commonCardList, "Common Card List")
        elif check == 3:
            printCards(uncommonCardList, "Uncommon Card List")
        elif check == 4:
            printCards(rareCardList, "Rare Card List")
        elif check == 5:
            printCards(mythicCardList, "Mythic Card List")
        elif check == 6:
            printCards(landCardList, "Land Card List")
        elif check == 7:
            printCards(allCardList, "All Card List")
        elif check == 0:
            print("\nThank you for playing!")
            quit()
        else:
            print("\nError: Unknown Code! Retry input number.")

if __name__ == "__main__":
    main()
